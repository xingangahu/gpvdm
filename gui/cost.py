#    General-purpose Photovoltaic Device Model - a drift diffusion base/Shockley-Read-Hall
#    model for 1st, 2nd and 3rd generation solar cells.
#    Copyright (C) 2012-2017 Roderick C. I. MacKenzie r.c.i.mackenzie at googlemail.com
#
#	https://www.gpvdm.com
#	Room B86 Coates, University Park, Nottingham, NG7 2RD, UK
#
#    This program is free software; you can redistribute it and/or modify
#    it under the terms of the GNU General Public License v2.0, as published by
#    the Free Software Foundation.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    You should have received a copy of the GNU General Public License along
#    with this program; if not, write to the Free Software Foundation, Inc.,
#    51 Franklin Street, Fifth Floor, Boston, MA 02110-1301 USA.


import os
from tab import tab_class
from icon_lib import QIcon_load

#qt
from PyQt5.QtCore import QSize, Qt 
from PyQt5.QtWidgets import QWidget,QVBoxLayout,QToolBar,QSizePolicy,QAction,QTableWidget,QAbstractItemView
from PyQt5.QtGui import QPainter,QIcon

#python modules
import webbrowser

from help import help_window

from epitaxy import epitaxy_get_layers
from epitaxy import epitaxy_get_width
from mesh import mesh_get_xlen
from mesh import mesh_get_zlen
from epitaxy import epitaxy_get_mat_file

work_book_enabled=False
try:
	from openpyxl import Workbook
	from openpyxl import load_workbook
	work_book_enabled=True
except:
	print("pyton3-openpyxl not found")

from gui_util import tab_add


from cal_path import get_materials_path

from inp import inp_get_token_value
from QWidgetSavePos import QWidgetSavePos

articles = []
mesh_articles = []

class cost(QWidgetSavePos):


	def callback_help(self):
		webbrowser.open('http://www.gpvdm.com/man/index.html')

	def __init__(self):
		QWidgetSavePos.__init__(self,"cost")
		self.setFixedSize(900, 600)
		self.setWindowIcon(QIcon_load("jv"))

		self.setWindowTitle(_("Cost and energy payback calculator (BETA - missing realistic data at the moment!!!)")) 
		

		self.main_vbox = QVBoxLayout()

		toolbar=QToolBar()
		toolbar.setIconSize(QSize(48, 48))

		self.play = QAction(QIcon_load("media-playback-start"), _("Re-calcualte"), self)
		self.play.triggered.connect(self.update)
		toolbar.addAction(self.play)
		
		spacer = QWidget()
		spacer.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
		toolbar.addWidget(spacer)


		self.help = QAction(QIcon_load("help"), _("Help"), self)
		self.help.triggered.connect(self.callback_help)
		toolbar.addAction(self.help)

		self.main_vbox.addWidget(toolbar)


		self.tab= QTableWidget()

		self.main_vbox.addWidget(self.tab)



		self.setLayout(self.main_vbox)

		self.update()

	def update(self):
		self.tab.clear()
		self.tab.setColumnCount(5)
		self.tab.setRowCount(0)
		self.tab.setSelectionBehavior(QAbstractItemView.SelectRows)
		self.tab.setHorizontalHeaderLabels([_("material"), _("Volume")+" (m^-3)", _("Mass")+" (kg)", _("Cost")+" ($)", _("Energy")+" (J)"])
		self.tab.setColumnWidth(1, 200)
		self.tab.setColumnWidth(2, 200)
		self.tab.setColumnWidth(3, 200)
		self.tab.setColumnWidth(4, 200)
		if work_book_enabled==False:
			print(_("pyton3-openpyxl not found"))
			return

		energy_tot=0.0
		cost_tot=0.0
		for i in range(0,epitaxy_get_layers()):
			
			volume=epitaxy_get_width(i)*1.0*1.0
			name=epitaxy_get_mat_file(i)
			xls_file_name=os.path.join(get_materials_path(),epitaxy_get_mat_file(i),"cost.xlsx")
			if os.path.isfile(xls_file_name):
				wb = load_workbook(xls_file_name)
				ws= wb.get_sheet_by_name("results")

				density = float(ws['B2'].value)
				mass=density*volume

				cost_per_kg = float(ws['B3'].value)
				cost=mass*cost_per_kg

				energy_per_kg = float(ws['B4'].value)
				energy=energy_per_kg*mass

				tab_add(self.tab,[name,str(volume),str(mass),str(cost),str(energy)])

				energy_tot=energy_tot+energy
				cost_tot=cost_tot+cost
		
		pce=inp_get_token_value("sim_info.dat", "#pce")
		payback_time=-1.0
		if pce!=None:
			pce=float(pce)
			gen_energy=1366.0*pce/100.0
			payback_time=energy_tot/gen_energy/60.0/60.0/24/365
		
		tab_add(self.tab,["sum","","",str(cost_tot),str(energy_tot)])
		tab_add(self.tab,["","","pay back time=",str(payback_time),"years"])




